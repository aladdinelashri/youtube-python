import PyPDF2
import openpyxl
import re

# Get the path to the PDF file
pdf_file_path = "C:\\activitytest.pdf"

# Create an Excel workbook
wb = openpyxl.Workbook()

# Create a sheet in the workbook
sheet = wb.active

# Open the PDF file in binary mode
with open(pdf_file_path, 'rb') as pdf_file:
    
    # Create a PDF reader object
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    
   
 # Iterate over the pages of the PDF file
    for page_num in range(len(pdf_reader.pages)):
       # Get the page object using the getPage() method
        page_obj = pdf_reader.pages [page_num]
        
        
        # Extract the text from the page
        text = page_obj.extract_text()
        
        # Pring page no which extracted and also the text extracted 
        print(f'Extracted text from page {page_num + 1}:')
        print(text)
   
   

    # Find all the YouTube video names and URL addresses
    video_names = re.findall(r"Watched (.*) on YouTube", text)
    video_urls = re.findall(r"https://www.youtube.com/watch\?v=(.*)", text)
    
    # Print how many video named and how many video url foud
    print(f'Found {len(video_names)} video names and {len(video_urls)} video URLs.')

    # Add the video names and URL addresses to the Excel sheet
    for i in range(len(video_names)):
        sheet.cell(row=i + 1, column=1).value = video_names[i]
        sheet.cell(row=i + 1, column=2).value = video_urls[i]

# Save the Excel workbook
wb.save("youtube_activity.xlsx")