import PyPDF2
import openpyxl

pdfFile = PyPDF2.PdfReader("C:\\activitytest.pdf", strict=False)

key = "/Annots"
uri = "/URI"
ank = "/A"
mylist = set()

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "YouTube Links"

row = 1
for page_no in range(len(pdfFile.pages)):
    page = pdfFile.pages[page_no]
    text = page.extract_text()
    print(f'Extracted text from page {page_no + 1}:')
    pageObject = page.get_object()
    if key in pageObject.keys():
        ann = pageObject[key]
        for a in ann:
            try:
                u = a.get_object()
                if uri in u[ank].keys():
                        link = u[ank][uri]
                        if link.startswith("https://www.youtube.com/watch?v="):
                            if link not in mylist:
                                mylist.add(link)
                                name = link.split("=")[1]
                                title = get_title(text)
                                ws.cell(row, 1).value = link
                                ws.cell(row, 2).value = name
                                ws.cell(row, 3).value = title
                                row += 1
            except KeyError:
                    pass

wb.save("youtube_links.xlsx")

print("The number of unique YouTube links is:", len(mylist))

def get_title(text):
    """Gets the title of the PDF document."""
    title = ""
    for page in text.split("\n"):
        if "Title:" in page:
            title = page.split(":")[1]
            break
    return title
